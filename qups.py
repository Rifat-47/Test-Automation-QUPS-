from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
import time
from datetime import datetime
from openpyxl import load_workbook

# Set Chrome options to modify browser settings
chrome_options = webdriver.ChromeOptions()

# Initialize the Chrome WebDriver with the modified options
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options) 

# Open Google in the Chrome browser
driver.get("https://www.google.com")

driver.find_element(by='xpath', value="//div[@id='SIvCob']/a").click()
time.sleep(2)

# Get the current day
today = datetime.now().strftime("%A")

# Load the workbook
workbook = load_workbook('qups.xlsx')
sheet = workbook[today]

# Column to be checked
column_letter = 'B'

# Find the highest & lowest row index of the column
min_row_index = 2
highest_row_index = sheet.max_row

# Find the search box element
search_box = driver.find_element(by='xpath', value="//textarea[@id='APjFqb']")

for index in range(min_row_index, highest_row_index+1):
    # getting the keyword to be searched 
    search_query = sheet[f'{column_letter}{index}'].value
    
    search_box.send_keys(search_query)
    time.sleep(2)

    # getting all the suggestions
    li_elements = driver.find_elements(by='xpath', value="//ul[@role='listbox'][1]/li")

    string_list = []
    for li_elem in li_elements:
        string_list.append(li_elem.find_element(by='xpath', value="div/div[2]/div/div/span").text)
    
    # getting longest and shortest string and update on spreadsheet
    longest = max(string_list, key=len)
    shortest = min(string_list, key=len)
    sheet[f'C{index}'] = longest
    sheet[f'D{index}'] = shortest

    search_box.clear()
    time.sleep(1)

# Save the workbook
workbook.save('qups.xlsx')
print("Data saved successfully")

time.sleep(5)

# Close the browser
driver.close()